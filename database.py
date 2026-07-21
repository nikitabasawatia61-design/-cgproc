import json
import sqlite3
from datetime import datetime
from pathlib import Path

from scraper.location import extract_area_city
from scraper.dates import is_tender_closed, parse_last_date

DB_PATH = Path(__file__).parent / "tenders.db"
DATA_JSON = Path(__file__).parent / "docs" / "data" / "tenders.json"


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS tenders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tender_no TEXT UNIQUE NOT NULL,
                name TEXT,
                department TEXT,
                amount TEXT,
                last_date TEXT,
                first_seen_at TEXT NOT NULL,
                last_updated_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_tenders_first_seen
            ON tenders(first_seen_at DESC)
        """)
        _ensure_area_city_column(conn)
        conn.commit()


def _ensure_area_city_column(conn):
    columns = {row[1] for row in conn.execute("PRAGMA table_info(tenders)")}
    if "area_city" not in columns:
        conn.execute("ALTER TABLE tenders ADD COLUMN area_city TEXT")


def get_existing_tender_numbers():
    with get_connection() as conn:
        rows = conn.execute("SELECT tender_no FROM tenders").fetchall()
    return {row["tender_no"] for row in rows}


def save_tender(data):
    now = datetime.now().isoformat(timespec="seconds")
    area_city = data.get("area_city") or extract_area_city(
        name=data.get("name", ""),
        department=data.get("department", ""),
    )
    with get_connection() as conn:
        _ensure_area_city_column(conn)
        conn.execute(
            """
            INSERT INTO tenders
                (tender_no, name, department, amount, last_date, area_city, first_seen_at, last_updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(tender_no) DO UPDATE SET
                name = excluded.name,
                department = excluded.department,
                amount = excluded.amount,
                last_date = excluded.last_date,
                area_city = excluded.area_city,
                last_updated_at = excluded.last_updated_at
            """,
            (
                data["tender_no"],
                data.get("name", ""),
                data.get("department", ""),
                data.get("amount", ""),
                data.get("last_date", ""),
                area_city,
                now,
                now,
            ),
        )
        conn.commit()


def get_all_tenders(search=None, new_only=False, include_closed=True, limit=500):
    query = "SELECT * FROM tenders WHERE 1=1"
    params = []

    if search:
        query += " AND (tender_no LIKE ? OR name LIKE ? OR department LIKE ? OR area_city LIKE ?)"
        term = f"%{search}%"
        params.extend([term, term, term, term])

    if new_only:
        today = datetime.now().strftime("%Y-%m-%d")
        query += " AND first_seen_at LIKE ?"
        params.append(f"{today}%")

    query += " ORDER BY first_seen_at DESC, CAST(tender_no AS INTEGER) DESC LIMIT ?"
    params.append(limit)

    with get_connection() as conn:
        rows = conn.execute(query, params).fetchall()

    tenders = [dict(row) for row in rows]
    if include_closed:
        return tenders
    return [t for t in tenders if not is_tender_closed(t.get("last_date"))]


def get_tender_by_number(tender_no):
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM tenders WHERE tender_no = ?",
            (tender_no,),
        ).fetchone()
    return dict(row) if row else None


def get_stats():
    with get_connection() as conn:
        rows = conn.execute("SELECT last_date, first_seen_at, last_updated_at FROM tenders").fetchall()

    active_rows = [row for row in rows if not is_tender_closed(row["last_date"])]
    closed_rows = [row for row in rows if is_tender_closed(row["last_date"])]
    today = datetime.now().strftime("%Y-%m-%d")
    new_today = sum(
        1 for row in active_rows
        if (row["first_seen_at"] or "").startswith(today)
    )
    last_scraped = max(
        (row["last_updated_at"] for row in active_rows if row["last_updated_at"]),
        default=None,
    )

    return {
        "total": len(active_rows),
        "closed": len(closed_rows),
        "new_today": new_today,
        "last_scraped": last_scraped,
    }


def remove_closed_tenders():
    """Keep closed tenders in the database for the Closed tab."""
    return 0


def remove_tenders_not_on_portal(portal_numbers):
    """Keep all tenders in the database; the UI splits open vs closed by due date."""
    return 0


def import_from_excel(excel_path):
    """One-time migration from existing Excel file."""
    from openpyxl import load_workbook

    if not Path(excel_path).exists():
        return 0

    wb = load_workbook(excel_path)
    ws = wb.active
    imported = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        save_tender({
            "tender_no": str(row[0]).strip(),
            "name": row[1] or "",
            "department": row[2] or "",
            "amount": row[3] or "",
            "last_date": row[4] or "",
        })
        imported += 1

    return imported


def import_from_json(json_path=None):
    """Load existing tenders from the GitHub Pages data file."""
    path = Path(json_path or DATA_JSON)
    if not path.exists():
        return 0

    with open(path, encoding="utf-8") as f:
        payload = json.load(f)

    imported = 0
    for tender in payload.get("tenders", []):
        if not tender.get("tender_no"):
            continue
        if import_tender_record(tender):
            imported += 1

    return imported


def import_tender_record(tender):
    first_seen = tender.get("first_seen_at") or datetime.now().isoformat(timespec="seconds")
    last_updated = tender.get("last_updated_at") or first_seen

    with get_connection() as conn:
        existing = conn.execute(
            "SELECT tender_no FROM tenders WHERE tender_no = ?",
            (tender["tender_no"],),
        ).fetchone()
        if existing:
            return False

        conn.execute(
            """
            INSERT INTO tenders
                (tender_no, name, department, amount, last_date, area_city, first_seen_at, last_updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                tender["tender_no"],
                tender.get("name", ""),
                tender.get("department", ""),
                tender.get("amount", ""),
                tender.get("last_date", ""),
                tender.get("area_city")
                or extract_area_city(tender.get("name", ""), tender.get("department", "")),
                first_seen,
                last_updated,
            ),
        )
        conn.commit()
    return True


def export_to_json(json_path=None):
    """Export open and closed tenders for GitHub Pages UI."""
    path = Path(json_path or DATA_JSON)
    path.parent.mkdir(parents=True, exist_ok=True)

    payload = {
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "stats": get_stats(),
        "tenders": get_all_tenders(limit=50000, include_closed=True),
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    return path


def backfill_area_city(force_all=False):
    """Fill or refresh area_city from description and department."""
    updated = 0
    with get_connection() as conn:
        _ensure_area_city_column(conn)
        rows = conn.execute(
            "SELECT tender_no, name, department, area_city FROM tenders"
        ).fetchall()

        for row in rows:
            area_city = extract_area_city(row["name"], row["department"])
            if not area_city:
                continue
            if not force_all and row["area_city"] == area_city:
                continue
            conn.execute(
                "UPDATE tenders SET area_city = ? WHERE tender_no = ?",
                (area_city, row["tender_no"]),
            )
            updated += 1
        conn.commit()
    return updated
